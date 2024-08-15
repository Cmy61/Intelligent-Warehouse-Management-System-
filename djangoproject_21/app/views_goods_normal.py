import shelve
from http.client import HTTPResponse
import xlrd
import os
import openpyxl
from django.utils.crypto import get_random_string
from django.http import  JsonResponse
from app.models import goods,store,shelves
from django.http import  JsonResponse
import xlrd
from django.shortcuts import render,redirect
from django.http import HttpResponse
from http.client import HTTPResponse

def GoodsStore(request):
    #return render(request,'store.html')
    return queryAllStore(request)

# 查询全部仓库信息
def queryAllStore(request):
    # 查询Store表的全部记录
    lstStore=store.objects.all()
    # 封装到传递参数中
    context=dict()
    context['lstStore']=lstStore
    # 页面响应跳转
    return render(request,'GoodStore_normal.html',context)

def showGoods(request,id):
    lstGoods=goods.objects.filter(goodsStore = id).all()
    # 封装到传递参数中
    context=dict()
    context['lstGoods']=lstGoods
    # 页面响应跳转
    return render(request,'goods_normal.html',context)

# 增加货物信息
def addGoods(request):
    if request.POST:
        # 接收请求数据
        goodsType=request.POST.get('addGoodsType',None)# 这里可以在前端设置有哪些选项
        goodsName=request.POST.get('addGoodsName',None)# 这里可以在前端设置有哪些选项
        goodsStore=request.POST.get('addGoodsStore',None)# 这里可以在前端设置有哪些选项
        # goods_lastStore=request.POST.get('goods_lastStore',None)# 这里可以在前端设置有哪些选项
        # goods_date=request.POST.get('goods_date',None)
        goodsAmount=request.POST.get('addGoodsAmount',None)
        goodsShelves=request.POST.get('addGoodsShelves',None)
        ulr='/app/gotoGoods_normal/goto/'+goodsStore
        goodsStore=store(goodsStore)
        
        # try:
            # 存储响应内容
        goods.objects.create(goodsType=goodsType,goodsName=goodsName,goodsStore=goodsStore,goodsShelves=goodsShelves,goodsAmount=goodsAmount)
       
            # goods.objects.create(goodsName=goodsName)
            # goods.objects.create(goodsStore=goodsStore)
            # goods.objects.create(goodsShelves=goodsShelves)
            # # goods.objects.create(goods_date=goods_date)
            # goods.objects.create(goodsAmount=goodsAmount)
            # 响应客户端
            
       
        return redirect(ulr)
            # return render(request,'addStore.html',{'message':'[OK].数据添加成功.'})

        # except:
        #     # 响应客户端
        #     return render(request,'goods.html',{'message':'[Error].数据添加失败.'})

    else:
        # 响应客户端
        return render(request,'goods_normal.html')

# 删除某物料信息
def  deleteGoods(request,id):
        # 调用删除函数
    Store=goods.objects.filter(goodsId = id).values('goodsStore')
    Store=list(Store)
    Store=Store[0]
    Store=str(Store['goodsStore'])
    print(Store)
    goods.objects.filter(goodsId = id).delete()
        # 响应客户端
    # print(request.get_full_path())
    
    url='/app/gotoGoods_normal/goto/'+Store
    print(url)
    return redirect(url)
# def  deleteStore(request,id):
#     # 调用删除函数
#     print(id)
#     store.objects.filter(storeId = id).delete()
#     # 响应客户端
#     return redirect('/app/index/')


# 更新物料信息
def updateGoods(request):
    if request.POST:
        # 获取新的内容
        
        goodsId=request.POST.get('updategoodsId',None)
        goodsType=request.POST.get('updategoodsType',None)
        goodsName=request.POST.get('updategoodsName',None)
        goodsStore=request.POST.get('updategoodsStore',None)
        goodsAmount=request.POST.get('updategoodsAmount',None)
        goodsShelves=request.POST.get('updategoodsShelves',None)
        # goodsShelves=shelves(goodsShelves)
        url='/app/gotoGoods_normal/goto/'+goodsStore
        print(goodsStore)
        goodsStore=store(goodsStore)
        
        # goods_lastStore=request.POST.get('goods_lastStore',None)
        # 处理客户端请求
        goods.objects.filter(goodsId = goodsId).update(goodsType=goodsType,goodsName=goodsName,goodsShelves=goodsShelves,goodsStore=goodsStore,goodsAmount=goodsAmount)
        # 响应客户端
        goodsStore=str(goodsStore)
        
        return redirect(url)

# 查询全部物料信息
# def queryAllGoods(request):
#     # 查询Store表的全部记录
#     lstGoods=goods.objects.all()
#     # 封装到传递参数中
#     context=dict()
#     context['lstStore']=lstGoods
#     # 页面响应跳转
#     return render(request,'#',context)


# 查询特定物料
def searchCertainGoods(request):
    print("searchCertainGoods2")
    searchName=request.POST.get('inquiryGoodsName',None)
    searchStore=request.POST.get('inquiryGoodsStore',None)
    searchType=request.POST.get('inquiryGoodsType',None)
    searchShelves=request.POST.get('inquiryGoodsShelves',None)
    
    context2=dict()
    print(searchType)
    print(searchName)
    searchStore=int(searchStore)
    if searchName=='' and searchShelves=='' and searchType=='':
        return redirect('/app/index2_normal/')
    else:
        context2=dict()
        print(searchName)
        print(searchStore)
        print(searchType)
        print(searchShelves)
        # if(searchLoca==''):
        #     searchLoca=None
        
        getId=[] 
        if(searchShelves==''):
            print("searchShelves")
            getId=goods.objects.filter(goodsName__icontains=searchName,goodsStore=searchStore)        
        else:
            print("searchShelves不是空")
            getId=goods.objects.filter(goodsName__icontains=searchName,goodsStore=searchStore,goodsType__icontains=searchType,goodsShelves=searchShelves)        
        print(getId)
        context2['lstGoods']=getId
        
        return render(request,'goodsSearch_normal.html',context2)

# 更新物料信息前返回数据
def show(request,id):
    # print(123)
    # print(id)
    # info=store.objects.get(storeId = id)
    # print(info)
    Name=goods.objects.filter(goodsId = id).values('goodsName')# 返回id对应的姓名
    Type=goods.objects.filter(goodsId = id).values('goodsType')
    Store=goods.objects.filter(goodsId = id).values('goodsStore')
    Amount=goods.objects.filter(goodsId = id).values('goodsAmount')
    Date=goods.objects.filter(goodsId = id).values('goodsDate')
    Shelves=goods.objects.filter(goodsId = id).values('goodsShelves')
    #因为这里得到的变量是queryset类型，先转为list类型，再取得list中的第一对象是dict类型，再从dict中取得所需内容
    Store=list(Store)
    Store=Store[0]
    Name=list(Name)
    Name=Name[0]
    Amount=list(Amount)
    Amount=Amount[0]
    Date=list(Date)
    Date=Date[0]
    Type=list(Type)
    Type=Type[0]
    Shelves=list(Shelves)
    Shelves=Shelves[0]
    #print(Loca)
    #print(type(Loca))
    #print(Loca['storeLoca'])
    data=dict()
    
    # 打包成dict传回去
    data['showgoodsName'] = Name['goodsName']
    data['showgoodsType']=Type['goodsType']
    data['showgoodsAmount']=Amount['goodsAmount']
    data['showgoodsStore']=str(Store['goodsStore'])
    # data['showgoodsDate']=Store['goodsDate']
    data['showgoodsId']=str(id)
    data['showgoodsShelves']=str(Shelves['goodsShelves'])
    #print(data['showstoreCap'])
    print(data)
    return JsonResponse(data)

#批量导入物料信息
def import_Gexcel(request):
    if request.method == "POST":
        print("POST")
        flag = False
        # try:
            #获取上传文件
        paper_file = request.FILES.get('file',None)
        excel_type = paper_file.name.split('.')[1]
        #判断文件类型
        print("FILE")
        if excel_type in ['xlsx','xls']:
            wb = xlrd.open_workbook(filename=None,file_contents=paper_file.read())
            # 根据地址读取工作表名(excel里是什么名字,这里就写什么名字)
            # sheet = wb.sheet_by_name("sheet1")
            wb_sheets = wb.sheets()
            for sheet in wb_sheets:
            # 在读取的过程中把第一行不做读取范围，然后进行循环读取数据并且进行保存，下面注释的都是数据表里自动生成的。
                for r in range(1, sheet.nrows):
                    print("for success")
                    goods.objects.create(goodsId=sheet.cell(r, 0).value, 
                                                goodsType=sheet.cell(r, 1).value, 
                                                goodsName=sheet.cell(r, 2).value,
                                                goodsAmount=sheet.cell(r, 3).value, 
                                                goodsShelves=sheet.cell(r, 4).value,
                                                goodsStore=store(sheet.cell(r, 5).value))
                        # goodsId = sheet.cell(r, 0).value
                        # goodsType = sheet.cell(r, 1).value
                        # goodsName = sheet.cell(r, 2).value 
                        # goodsStore = sheet.cell(r, 3).value
                        # goodsDate = sheet.cell(r, 4).value
                        # goodsAmount = sheet.cell(r, 5).value
                        # goods_lastStore = sheet.cell(r, 5).value
                        # save_excel = goods.objects.create(goodsId=goodsId,goodsType=goodsType,goodsName=goodsName,goodsStore=goodsStore,goodsDate=goodsDate,goodsAmount=goodsAmount,goods_lastStore=goods_lastStore)
            flag = True
            return HttpResponse(flag)
        # except Exception as ex:
        #     return HttpResponse(flag)


def export_goods(request):
    media_root = '/media/'
    objs = goods.objects.all().values()
    stores = list(objs)
    excel_name = get_random_string() + '.xlsx'
    # 准备写入的路径
    path = os.path.join(media_root, excel_name)
    # 写入Excel
    to_Gexcel(stores, path)
    # 返回
    return JsonResponse(
        {"data": "/media/" + excel_name, "message": "操作成功"}
    )


def to_Gexcel(data: list, path: str):
    # 实例化一个workbook
    workbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = workbook.active
    # 为sheet命名
    sheet.title = 'Goods'
    # 准备keys
    keys = data[0].keys()
    # 写标题的第一行
    title_list = ['goodsID',  'goodsType','goodsName', 'goodsStore', 'goods_lastStore']
    for ix, item in enumerate(title_list):
        sheet.cell(row=1, column=ix + 1, value=item)
    # 准备写入数据
    for ix, item in enumerate(data):
        # 遍历每个元素
        for k, v in enumerate(keys):
            # 这里要从第二列开始
            sheet.cell(row=ix + 2, column=k + 1, value=str(item[v]))
    # 写入到文件
    workbook.save(path)