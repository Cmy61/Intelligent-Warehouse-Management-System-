import os

import xlwt
from django.shortcuts import render, redirect
from xlsxwriter import Workbook
from app.models import store
from app.views_notice import send_notifications
from django.http import JsonResponse
import xlrd
import openpyxl
from django.utils.crypto import get_random_string
from django.http import HttpResponse
from djangoproject.settings import BASE_DIR


def index(request):
    # return render(request,'store.html')
    return queryAllStore(request)


# 去store1.html
def gotoStore1(request):
    return render(request, 'store1.html')


# 增加仓库信息
def addStore(request):
    print('USER', request.user.userClass)
    if request.POST:
        # 接收请求数据
        storeName = request.POST.get('storeName', None)
        storeType = request.POST.get('storeType', None)  # 这里可以在前端设置有哪些选项
        storeLoca = request.POST.get('storeLoca', None)  # 这里可以在前端设置有哪些选项
        # storeExist = request.POST.get('storeExist', 0)  # 已使用容积
        storeCap = request.POST.get('storeCap', 0)

        try:
            # 存储响应内容
            store.objects.create(storeName=storeName, storeType=storeType, storeLoca=storeLoca,
                                 storeCap=storeCap)
            # [TEST] 通知公告
            send_notifications(sender=request.user, verb='添加仓库 ' + storeName, receiver='仓库管理员')
            print(request.user.notifications.unread())
            # 响应客户端
            return redirect('/app/index/')
            # return render(request,'store.html')

        except:
            #     # 响应客户端
            return redirect('/app/index/')

    else:
        # 响应客户端
        return render(request, 'store.html')


# # 删除某条仓库信息
# def  deleteStore(request,id):
#     # 调用删除函数
#     print(id)
#     store.objects.filter(storeId = id).delete()
#     # 响应客户端
#     return redirect('/app/index/')

# 删除某条仓库信息,含公告
def deleteStore(request, id):
    # 调用删除函数
    to_delete = store.objects.get(storeId=id).storeName
    store.objects.filter(storeId=id).delete()
    print('[TEST] delete store 通知公告')
    send_notifications(sender=request.user, verb='删除仓库 ' + to_delete, receiver='仓库管理员', )
    print(request.user.notifications.unread())
    # 响应客户端
    return redirect('/app/index/')


# 更新仓库信息
def updateStore(request):
    if request.POST:
        # 获取新的内容
        storeId = request.POST.get('updatestoreId', None)
        storeName = request.POST.get('updatestoreName', None)
        storeType = request.POST.get('updatestoreType', None)
        storeCap = request.POST.get('updatestoreCap', None)
        storeLoca = request.POST.get('updatestoreLoca', None)
        # storeExist = request.POST.get('updatestoreExist', None)
        # 处理客户端请求
        print(storeName)
        store.objects.filter(storeId=storeId).update(storeName=storeName, storeLoca=storeLoca,
                                                     storeType=storeType, storeCap=storeCap)
        # [TEST] 通知公告
        send_notifications(sender=request.user, verb='更新仓库 ' + storeName, receiver='仓库管理员', )
        # 响应客户端
        return queryAllStore(request)


# 查询全部仓库信息
def queryAllStore(request):
    # 查询Store表的全部记录
    lstStore = store.objects.all()
    # 封装到传递参数中
    context = dict()
    context['lstStore'] = lstStore
    # 页面响应跳转
    return render(request, 'store.html', context)


# 查询特定名称仓库
def searchCertainStore(request):
    searchName = request.POST.get('inquirystoreName', '')
    searchLoca = request.POST.get('inquirystoreLoca', '')
    searchType = request.POST.get('inquirystoreType', '')
    context2 = dict()
    print(searchName)
    print(searchLoca)
    print(searchType)
    # if(searchLoca==''):
    #     searchLoca=None

    getId = []
    getId = store.objects.filter(storeName__icontains=searchName, storeLoca__icontains=searchLoca,
                                 storeType__icontains=searchType)
    print(getId)
    context2['lstStore'] = getId
    return render(request, 'storeSearch.html', context2)


# 得到某条仓库信息
# def  getStore(request,id):
#     print(id)
#     #students.objects().filter(name = ‘张三').values('id'), 只返回名为张三的学生的id,不返回其他属性了。
#     Name=store.objects().filter(storeId = id).values('storeName')
#     # 响应客户端
#     print(Name)
#     return redirect('/app/index/')


#   sex = models.User.objects.get(UserID=1)
#     print(sex.UserGender)

# def updateStoreData(request):
#     id=request.POST.get('storeId')
#     Name=request.POST.get('storeName')
#     Loca=request.POST.get('storeLoca')
#     Cap=request.POST.get('storeCap')
#     Type=request.POST.get('storeType')
#     id=int(id)
#     store.objects.filter(storeId=id).update(storeName=Name,is_active=True)

# User.objects.filter(id=1).update(username='nick',is_active=True)
def show(request, id):
    # info=store.objects.get(storeId = id)
    # print(info)
    Name = store.objects.filter(storeId=id).values('storeName')  # 返回id对应的姓名
    Type = store.objects.filter(storeId=id).values('storeType')
    Cap = store.objects.filter(storeId=id).values('storeCap')
    # Exist = store.objects.filter(storeId=id).values('storeExist')
    Loca = store.objects.filter(storeId=id).values('storeLoca')
    # 因为这里得到的变量是queryset类型，先转为list类型，再取得list中的第一对象是dict类型，再从dict中取得所需内容
    Loca = list(Loca)
    Loca = Loca[0]
    Name = list(Name)
    Name = Name[0]
    # Exist = list(Exist)
    # Exist = Exist[0]
    Cap = list(Cap)
    Cap = Cap[0]
    Type = list(Type)
    Type = Type[0]
    # print(Loca)
    # print(type(Loca))
    # print(Loca['storeLoca'])
    data = dict()
    # 打包成dict传回去
    data['showstoreName'] = Name['storeName']
    data['showstoreType'] = Type['storeType']
    data['showstoreLoca'] = Loca['storeLoca']
    # data['showstoreExist'] = Exist['storeExist']
    data['showstoreCap'] = Cap['storeCap']
    data['showstoreId'] = id
    # print(data['showstoreCap'])
    return JsonResponse(data)


# 批量导入仓库信息
def import_Sexcel(request):
    if request.method == 'POST':
        flag = False
        # print("POST提交成功")
        try:
            # 获取上传文件
            doc = request.FILES.get('file', None)
            # msg = 'success' if doc else 'false'
            # print('doc prepared', msg)
            doc_type = doc.name.split('.')[1]
            # 判断文件类型
            if doc_type in ['xlsx', 'xls']:
                wb = xlrd.open_workbook(filename=None, file_contents=doc.read())
                # 根据地址读取工作表名
                wb_sheets = wb.sheets()
                store_list = []
                # 读取所有sheet
                for sheet in wb_sheets:
                    # 在读取的过程中把第一行不做读取范围，然后进行循环读取数据并且进行保存，下面注释的都是数据表里自动生成的。
                    for r in range(1, sheet.nrows):
                        store_list.append(sheet.cell(r, 1).value)
                        store.objects.create(storeId=sheet.cell(r, 0).value,
                                             storeName=sheet.cell(r, 1).value,
                                             storeType=sheet.cell(r, 2).value,
                                             storeLoca=sheet.cell(r, 3).value,
                                             storeCap=sheet.cell(r, 4).value)
                # [TEST] 通知公告
                send_notifications(sender=request.user, verb='导入仓库 ' + str(store_list), receiver='仓库管理员', )
                flag = True
                return HttpResponse(flag)
        except Exception as ex:
            return HttpResponse(flag)


def export_store(request):
    # response内容类型为excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    try:
        # media_root = os.path.join(BASE_DIR + '/media')
        # # 判断media文件夹是否存在
        # if not os.path.exists(media_root):
        #     os.makedirs(media_root)
        wb = Workbook(response, {'in_memory': True})
        ws = wb.add_worksheet('Stores')
        # sheet第一行
        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        excel_name = get_random_string(18)
        print(excel_name)
        # 下载文件名
        response['Content-Disposition'] = 'attachment; filename={0}.xlsx'.format(excel_name)
        # 准备写入的路径
        # path = os.path.join(media_root, excel_name)
        # 写入Excel
        # store_to_excel(stores, path)
        # 实例化一个workbook
        # workbook = openpyxl.Workbook()
        # # 激活一个sheet
        # sheet = workbook.active
        # # 为sheet命名
        # sheet.title = 'Store'
        # # 准备keys
        # keys = stores[0].keys()
        # print(keys)
        # sheet第一行标题
        columns = ['storeID', 'storeName', 'storeType', 'storeLoca', 'storeCap']
        # for ix, item in enumerate(title_list):
        #     sheet.cell(row=1, column=ix + 1, value=item)
        # print(stores[0])
        # # 准备写入数据
        # for ix, item in enumerate(stores[0]):
        #     # 遍历每个元素
        #     for k, v in enumerate(keys):
        #         # 这里要从第二行开始
        #         print(k, v)
        #         # sheet.cell(row=ix + 2, column=k + 1, value=str(item[v]))
        # print(sheet)
        # 写入到文件
        title_font = wb.add_format({'bold': True, 'font_color': 'violet'})
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], title_font)
                     # , font_style)

        # sheet其他行
        # font_style = xlwt.XFStyle()

        rows = store.objects.all().values_list('storeId', 'storeName', 'storeType',
                                               'storeLoca', 'storeCap')
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num])
                         # , font_style)
        wb.close()
        # 返回
        return response
    except Exception:
        return JsonResponse(
            {"message": "File export failed!"}
        )


def store_to_excel(data: list, path: str):
    # 实例化一个workbook
    workbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = workbook.active
    # 为sheet命名
    sheet.title = 'Store'
    # 准备keys
    keys = data[0].keys()
    # 写标题的第一行
    title_list = ['storeID', 'storeName', 'storeType', 'storeLoca', 'storeCap']
    for ix, item in enumerate(title_list):
        sheet.cell(row=1, column=ix + 1, value=item)
    # 准备写入数据
    for ix, item in enumerate(data):
        # 遍历每个元素
        for k, v in enumerate(keys):
            # 这里要从第二列开始
            sheet.cell(row=ix + 2, column=k + 1, value=str(item[v]))
    # 写入到文件
    workbook.save(path)
